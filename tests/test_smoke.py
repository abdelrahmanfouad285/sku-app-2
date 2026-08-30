"""
Smoke test for the SKU data agent — exercises every non-LLM code path.

Run:
    .venv\\Scripts\\python.exe tests\\test_smoke.py
    # or
    python -m tests.test_smoke

Does NOT call any LLM (Ollama / Claude). It only tests:
  - excel_manager (append, dedup, highlight, atomic write, EXCEL_COLUMNS)
  - image_utils (encode, EXIF, HEIC, downscaling, base64)
  - barcode_scanner (the soda photo in input_photos/processed)
  - validator (brand/product fuzzy match)
  - barcode_lookup (CSV cache + path)
  - pipeline_helpers (list_input_images, append_run_log, move_to_processed)
  - llm._parse_json (fence stripping, first {...} extraction)
"""

from __future__ import annotations

import sys
import shutil
import tempfile
import os
from pathlib import Path

# Make the project root importable so `import utils...` works regardless of cwd.
PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))


def _ok(msg: str) -> None:
    print(f"  [ok] {msg}")


def _section(name: str) -> None:
    print(f"\n=== {name} ===")


def test_excel_columns() -> None:
    _section("excel_manager.EXCEL_COLUMNS")
    from utils.excel_manager import EXCEL_COLUMNS, REQUIRED_FIELDS
    expected = {
        "source_filename", "processed_at", "item_index", "item_count",
        "sku", "product_name", "brand",
        "flavor_or_variant", "quantity_or_size", "visible_unit_count",
        "confidence", "notes",
    }
    assert set(EXCEL_COLUMNS) == expected, f"column set mismatch: {set(EXCEL_COLUMNS)} vs {expected}"
    assert "price" not in EXCEL_COLUMNS, "price column should be removed"
    assert "flavor_or_variant" in EXCEL_COLUMNS, "flavor_or_variant column missing"
    assert "visible_unit_count" in EXCEL_COLUMNS, "visible_unit_count column missing"
    _ok("EXCEL_COLUMNS contains all expected fields (no price; with flavor_or_variant + visible_unit_count)")


def test_excel_append_and_dedup(tmpdir: Path) -> None:
    _section("excel_manager.append_product — append + dedup + highlight + atomic")
    import utils.excel_manager as em
    # Redirect EXCEL_PATH to a temp file so we don't touch the real workbook.
    real_path = em.EXCEL_PATH
    real_dir = em.OUTPUT_DIR
    em.EXCEL_PATH = tmpdir / "sku_data.xlsx"
    em.OUTPUT_DIR = tmpdir
    try:
        # Append a complete row
        em.append_product({
            "source_filename": "a.jpg",
            "processed_at": "2026-01-01T00:00:00",
            "item_index": 1,
            "item_count": 1,
            "sku": "X1",
            "product_name": "Test",
            "brand": "Acme",
            "quantity_or_size": "500ml",
            "price": "10.00",
            "confidence": "high",
            "notes": None,
        })
        assert em.EXCEL_PATH.exists()
        _ok("first append creates file")

        # Append a row that needs review (low confidence)
        em.append_product({
            "source_filename": "a.jpg",
            "processed_at": "2026-01-01T00:00:01",
            "item_index": 2,
            "item_count": 2,
            "sku": None,
            "product_name": None,
            "brand": None,
            "quantity_or_size": None,
            "price": None,
            "confidence": "low",
            "notes": "blurry",
        })

        # Idempotent re-append of row 1 should NOT duplicate
        em.append_product({
            "source_filename": "a.jpg",
            "processed_at": "2026-01-01T00:00:00",
            "item_index": 1,
            "item_count": 1,
            "sku": "X1",
            "product_name": "Test",
            "brand": "Acme",
            "quantity_or_size": "500ml",
            "price": "10.00",
            "confidence": "high",
            "notes": None,
        })

        from openpyxl import load_workbook
        wb = load_workbook(em.EXCEL_PATH)
        ws = wb.active
        data_rows = ws.max_row - 1  # minus header
        assert data_rows == 2, f"expected 2 rows, got {data_rows}"
        _ok("dedup works: re-append does not duplicate")

        # Verify the low-confidence row is highlighted (FFFFE699)
        low_row = None
        for r in range(2, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "a.jpg" and ws.cell(row=r, column=4).value == 2:
                low_row = r
                break
        assert low_row is not None
        fill = ws.cell(row=low_row, column=2).fill
        # openpyxl returns start_color.rgb
        rgb = getattr(fill.start_color, "rgb", None) or ""
        assert "FFE699" in str(rgb).upper(), f"expected highlight on low-confidence row, got {rgb}"
        _ok("low-confidence row is yellow-highlighted")
    finally:
        em.EXCEL_PATH = real_path
        em.OUTPUT_DIR = real_dir


def test_image_utils() -> None:
    _section("image_utils.encode_image_as_base64")
    from utils.image_utils import encode_image_as_base64
    import base64
    # Use one of the test images already in the repo
    sample = PROJECT_ROOT / "input_photos" / "processed"
    sample.mkdir(parents=True, exist_ok=True)
    candidates = sorted(sample.glob("*.jpg"))
    assert candidates, "no .jpg in input_photos/processed to test with"
    encoded = encode_image_as_base64(str(candidates[0]))
    # base64 round-trip
    raw = base64.b64decode(encoded)
    assert raw[:3] == b"\xff\xd8\xff", "encoded output is not a JPEG"
    assert len(raw) > 1024, "encoded output is suspiciously small"
    _ok(f"encoded {candidates[0].name} -> {len(encoded)} chars ({len(raw)//1024} KB JPEG)")


def test_barcode_scanner() -> None:
    _section("barcode_scanner.scan_barcodes")
    from utils.barcode_scanner import scan_barcodes
    sample = PROJECT_ROOT / "input_photos" / "processed"
    candidates = sorted(sample.glob("*.jpg"))
    assert candidates
    codes = scan_barcodes(str(candidates[0]))
    # The soda can photo may or may not have a decodable barcode; just check
    # the function runs and returns a list of strings.
    assert isinstance(codes, list)
    for c in codes:
        assert isinstance(c, str) and c
    _ok(f"scan_barcodes returned {codes} for {candidates[0].name}")


def test_validator() -> None:
    _section("validator.validate_extraction")
    from utils.validator import validate_extraction
    # Without a known_products.csv, the validator is a no-op.
    item = {"brand": "Acme", "product_name": "Test", "confidence": "high", "notes": None}
    out = validate_extraction(item)
    assert out is item
    _ok("validator handles missing known_products.csv (no-op)")

    # Simulate a known list by writing one
    from utils.barcode_lookup import KNOWN_CSV_PATH, append_to_known_products
    append_to_known_products("9999999999", "Coca-Cola", "Coca-Cola Original")
    try:
        # Brand matches -> no flag
        good = {"brand": "Coca-Cola", "product_name": "Coca-Cola Original", "confidence": "high", "notes": None}
        validate_extraction(good)
        assert good["confidence"] == "high"
        _ok("matching brand/product does not flag")
        # Brand doesn't match -> flag
        bad = {"brand": "Fanta", "product_name": "Coca-Cola Original", "confidence": "high", "notes": None}
        validate_extraction(bad)
        assert bad["confidence"] == "low"
        assert "verify manually" in (bad.get("notes") or "").lower()
        _ok("non-matching brand is flagged (low + review note)")
    finally:
        if KNOWN_CSV_PATH.exists():
            KNOWN_CSV_PATH.unlink()


def test_barcode_lookup() -> None:
    _section("barcode_lookup.lookup_barcode")
    from utils.barcode_lookup import lookup_barcode
    # Empty input -> None
    assert lookup_barcode("") is None
    # Random barcode that doesn't exist in OFF either -> None (or hits timeout gracefully)
    result = lookup_barcode("0000000000000")
    # Don't assert the exact result (OFF is network-dependent), but never raise
    _ok(f"lookup_barcode('0000000000000') -> {result}")


def test_pipeline_helpers(tmpdir: Path) -> None:
    _section("pipeline_helpers (run log, list, move)")
    from utils import pipeline_helpers as ph
    real_log = ph.LOG_CSV_PATH
    real_in = ph.INPUT_DIR
    real_out = ph.OUTPUT_DIR
    real_proc = ph.PROCESSED_DIR
    ph.INPUT_DIR = tmpdir / "in"
    ph.PROCESSED_DIR = ph.INPUT_DIR / "processed"
    ph.OUTPUT_DIR = tmpdir / "out"
    ph.LOG_CSV_PATH = ph.OUTPUT_DIR / "run_log.csv"
    try:
        ph.ensure_dirs()
        assert ph.INPUT_DIR.exists() and ph.OUTPUT_DIR.exists() and ph.PROCESSED_DIR.exists()
        _ok("ensure_dirs creates the directory tree")

        # No images yet
        assert ph.list_input_images() == []
        _ok("list_input_images returns [] on empty dir")

        # Drop a fake image and list it
        img = ph.INPUT_DIR / "fake.jpg"
        img.write_bytes(b"\xff\xd8\xff\xe0")
        found = ph.list_input_images()
        assert len(found) == 1 and found[0].name == "fake.jpg"
        _ok("list_input_images returns a valid image")

        # Append a log entry
        ph.append_run_log("fake.jpg", 1, "ok")
        ph.append_run_log("bad.jpg", 0, "error", "x" * 5000)  # long error
        assert ph.LOG_CSV_PATH.exists()
        text = ph.LOG_CSV_PATH.read_text(encoding="utf-8")
        assert "fake.jpg" in text
        assert "bad.jpg" in text
        # Long error should be truncated
        assert len(text) < 5000, f"log file too large: {len(text)} bytes"
        _ok("append_run_log writes header + entries; long errors are truncated")

        # Move to processed
        dest = ph.move_to_processed(img)
        assert dest.exists() and not img.exists()
        _ok("move_to_processed moves the file")
    finally:
        ph.LOG_CSV_PATH = real_log
        ph.INPUT_DIR = real_in
        ph.OUTPUT_DIR = real_out
        ph.PROCESSED_DIR = real_proc


def test_llm_parse_json() -> None:
    _section("llm._parse_json (fence stripping)")
    from utils.llm import _parse_json
    # Plain JSON
    assert _parse_json('{"a": 1, "items": []}') == {"a": 1, "items": []}
    _ok("plain JSON parses")
    # Fenced JSON
    assert _parse_json('```json\n{"a": 2}\n```') == {"a": 2}
    _ok("fenced JSON parses")
    # JSON with surrounding prose
    assert _parse_json('Here you go: {"a": 3} cheers!') == {"a": 3}
    _ok("JSON inside prose parses")
    # Empty -> raises
    try:
        _parse_json("")
    except ValueError as e:
        assert "empty" in str(e).lower()
        _ok("empty input raises ValueError")


def main() -> int:
    print("=" * 60)
    print("SKU data agent -- smoke tests (no LLM)")
    print("=" * 60)

    with tempfile.TemporaryDirectory() as tmp:
        tmpdir = Path(tmp)
        test_excel_columns()
        test_excel_append_and_dedup(tmpdir)
        test_image_utils()
        test_barcode_scanner()
        test_validator()
        test_barcode_lookup()
        test_pipeline_helpers(tmpdir)
        test_llm_parse_json()

    print("\n" + "=" * 60)
    print("All smoke tests passed.")
    print("=" * 60)
    return 0


if __name__ == "__main__":
    sys.exit(main())

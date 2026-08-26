"""Excel I/O for SKU data rows.

Appends a single product row (or a DataFrame) to ./output/sku_data.xlsx,
creating the file with headers if it doesn't exist. Highlights rows that
need human review (low confidence OR missing any required field) in yellow.

Writes are atomic: a temp file is written first, then renamed over the
target, so a crash mid-write can never leave a half-written workbook on
disk. When the same (source_filename, item_index) already exists the
row is updated in place rather than duplicated, so retries are safe.

Self-healing: if an existing workbook has a header that doesn't match
EXCEL_COLUMNS (missing `price`, extra `Unnamed: N`, columns in a different
order), the file is rebuilt with the canonical header and existing rows
are remapped to the closest known column. This protects against a
schema-drift row shift that would otherwise silently corrupt every
subsequent append.
"""

from __future__ import annotations

import difflib
import os
import shutil
import tempfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


PROJECT_ROOT = Path(__file__).resolve().parent.parent
OUTPUT_DIR = PROJECT_ROOT / "output"
EXCEL_PATH = OUTPUT_DIR / "sku_data.xlsx"

EXCEL_COLUMNS = [
    "source_filename",
    "processed_at",
    "item_index",
    "item_count",
    "sku",
    "product_name",
    "brand",
    "flavor_or_variant",
    "quantity_or_size",
    "visible_unit_count",
    "confidence",
    "notes",
]

# Fields that must be non-empty for a row to be considered "complete".
REQUIRED_FIELDS = (
    "sku",
    "product_name",
    "brand",
    "quantity_or_size",
)

HIGHLIGHT_FILL = PatternFill(start_color="FFFFE699", end_color="FFFFE699", fill_type="solid")
HEADER_FILL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
HEADER_FONT = Font(bold=True)

# Aliases that older or LLM-generated headers might use instead of the
# canonical column names. Used by the self-heal pass to recover rows
# written under a different schema.
_HEADER_ALIASES: dict[str, list[str]] = {
    "source_filename": ["source_filename", "filename", "file", "image", "source"],
    "processed_at":    ["processed_at", "timestamp", "date", "time", "processed"],
    "item_index":      ["item_index", "index", "idx", "item #", "item"],
    "item_count":      ["item_count", "count", "total", "n_items"],
    "sku":             ["sku", "sku_number", "sku code", "item_code"],
    "barcode_number":  ["barcode_number", "barcode", "upc", "ean", "gtin"],
    "product_name":    ["product_name", "name", "title", "product", "description"],
    "brand":           ["brand", "manufacturer", "maker"],
    "quantity_or_size": ["quantity_or_size", "size", "quantity", "qty", "volume", "weight"],
    "flavor_or_variant": ["flavor_or_variant", "flavor", "variant", "taste"],
    "visible_unit_count": ["visible_unit_count", "unit_count", "units", "count", "visible_count"],
    "confidence":      ["confidence", "score", "certainty"],
    "notes":           ["notes", "note", "comment", "comments", "remarks"],
}


def _apply_header_style(ws) -> None:
    for col_idx, _ in enumerate(EXCEL_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")


def _apply_column_widths(ws) -> None:
    for col_idx, header in enumerate(EXCEL_COLUMNS, start=1):
        letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[letter].width = max(14, min(40, len(str(header)) + 4))


def _needs_review(row: dict) -> bool:
    """Highlight rule shared by all writers."""
    conf = (row.get("confidence") or "")
    if not isinstance(conf, str) or conf.lower() == "low":
        return True
    for f in REQUIRED_FIELDS:
        v = row.get(f)
        if v is None or v == "" or (isinstance(v, float) and pd.isna(v)):
            return True
    return False


def _load_or_create() -> tuple[Workbook, "Worksheet"]:
    """Open the existing workbook (or create a fresh one with headers).

    If the existing header doesn't match EXCEL_COLUMNS, self-heal by
    rebuilding the sheet under the canonical schema and remapping each
    existing row's cells to the closest known column.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        if not _header_is_canonical(ws):
            _self_heal_workbook(wb, ws)
        return wb, ws
    wb = Workbook()
    ws = wb.active
    ws.title = "SKU Data"
    ws.append(EXCEL_COLUMNS)
    _apply_header_style(ws)
    return wb, ws


def _header_is_canonical(ws) -> bool:
    """True if row 1 of `ws` exactly matches EXCEL_COLUMNS (and only those)."""
    if ws.max_column < 1:
        return False
    actual = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    return [str(a) if a is not None else "" for a in actual] == list(EXCEL_COLUMNS)


def _map_existing_header_to_canonical(ws) -> dict[int, int]:
    """Return {existing_col_index_1based: canonical_col_index_1based}.

    For each existing column header in `ws`, find the closest match in
    EXCEL_COLUMNS using aliases then difflib. Unmatched columns (e.g. the
    old phantom `Unnamed: N` or a `None` header) are dropped.
    """
    mapping: dict[int, int] = {}
    used: set[int] = set()
    for c in range(1, ws.max_column + 1):
        raw = ws.cell(row=1, column=c).value
        header = (str(raw).strip() if raw is not None else "").lower()
        if not header:
            continue  # drop blank / None headers
        canonical_idx = None
        # 1. exact match (case-insensitive)
        for ci, name in enumerate(EXCEL_COLUMNS, start=1):
            if name == header and ci not in used:
                canonical_idx = ci
                break
        # 2. alias match
        if canonical_idx is None:
            for ci, name in enumerate(EXCEL_COLUMNS, start=1):
                if ci in used:
                    continue
                if header in _HEADER_ALIASES.get(name, []):
                    canonical_idx = ci
                    break
        # 3. fuzzy match
        if canonical_idx is None:
            best_ratio = 0.0
            for ci, name in enumerate(EXCEL_COLUMNS, start=1):
                if ci in used:
                    continue
                ratio = difflib.SequenceMatcher(None, header, name).ratio()
                if ratio > best_ratio and ratio >= 0.7:
                    best_ratio = ratio
                    canonical_idx = ci
        if canonical_idx is not None:
            mapping[c] = canonical_idx
            used.add(canonical_idx)
    return mapping


def _self_heal_workbook(wb: Workbook, ws) -> None:
    """Rebuild `ws` with the canonical header and remap every existing row.

    Writes to a timestamped backup first so the user can recover if the
    remap miscategorises something. Logs to stdout so the Streamlit log
    panel can show the heal event.
    """
    backup = OUTPUT_DIR / f"sku_data.backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    try:
        shutil.copy2(EXCEL_PATH, backup)
        print(f"[excel] Header drift detected; backup saved to {backup.name}", flush=True)
    except Exception as exc:
        print(f"[excel] Could not write backup before self-heal: {exc}", flush=True)

    old_max_row = ws.max_row
    mapping = _map_existing_header_to_canonical(ws)

    # Snapshot existing data values (only mapped columns).
    old_data: list[list] = []
    for r in range(2, old_max_row + 1):
        row_vals: list = [None] * len(EXCEL_COLUMNS)
        for old_c, new_c in mapping.items():
            v = ws.cell(row=r, column=old_c).value
            row_vals[new_c - 1] = v
        old_data.append(row_vals)

    # Wipe the sheet and re-emit canonical header + remapped data.
    wb.remove(ws)
    new_ws = wb.create_sheet("SKU Data", 0)
    new_ws.append(EXCEL_COLUMNS)
    _apply_header_style(new_ws)
    for row_vals in old_data:
        new_ws.append(row_vals)
    _apply_column_widths(new_ws)
    _save_atomic(wb)
    print(
        f"[excel] Self-heal: remapped {len(mapping)}/{ws.max_column} old columns "
        f"to canonical schema, {len(old_data)} data row(s) preserved.",
        flush=True,
    )


def _save_atomic(wb: Workbook) -> None:
    """Write to a temp file in the same directory then rename. Prevents
    half-written workbooks if the process is killed mid-save."""
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    fd, tmp = tempfile.mkstemp(
        prefix=".sku_data_", suffix=".xlsx.tmp", dir=str(OUTPUT_DIR)
    )
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, EXCEL_PATH)  # atomic on Windows + POSIX
    except Exception:
        # Best-effort cleanup; never let a stray temp file linger.
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise


def _find_existing_row(ws, source_filename: str, item_index: int) -> int | None:
    """Return 1-based row index of (source_filename, item_index) in ws, or None."""
    if ws.max_row < 2:
        return None
    # Build column lookup from header row so we don't hard-code positions.
    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
    fn_col = header.get("source_filename")
    idx_col = header.get("item_index")
    if fn_col is None or idx_col is None:
        return None
    for r in range(2, ws.max_row + 1):
        fn = ws.cell(row=r, column=fn_col).value
        ix = ws.cell(row=r, column=idx_col).value
        if fn == source_filename and int(ix or 0) == int(item_index):
            return r
    return None


def append_product(row: dict) -> bool:
    """
    Append one product row to the Excel file. If a row with the same
    (source_filename, item_index) already exists it is updated in place
    — so retries after a partial failure are safe.

    Returns True if a new row was added, False if an existing row was
    updated.
    """
    wb, ws = _load_or_create()

    fn = row.get("source_filename") or ""
    try:
        idx = int(row.get("item_index") or 0)
    except (TypeError, ValueError):
        idx = 0

    existing = _find_existing_row(ws, fn, idx) if fn and idx > 0 else None

    values = [row.get(col) for col in EXCEL_COLUMNS]
    if existing is not None:
        for col_idx, v in enumerate(values, start=1):
            ws.cell(row=existing, column=col_idx).value = v
        excel_row = existing
    else:
        ws.append(values)
        excel_row = ws.max_row

    if _needs_review(row):
        for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
            ws.cell(row=excel_row, column=col_idx).fill = HIGHLIGHT_FILL
    else:
        # Clear any prior highlight on this row.
        for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
            ws.cell(row=excel_row, column=col_idx).fill = PatternFill(fill_type=None)

    _apply_column_widths(ws)
    _save_atomic(wb)
    return existing is None


def append_rows_to_excel(df: pd.DataFrame) -> None:
    """
    Append all rows from a pandas DataFrame to ./output/sku_data.xlsx.
    Creates the file with headers if it doesn't exist. Highlights rows
    that have low confidence or any null in a required field.

    Existing rows with the same (source_filename, item_index) are
    updated in place rather than duplicated.
    """
    wb, ws = _load_or_create()
    df = df.copy()
    for col in EXCEL_COLUMNS:
        if col not in df.columns:
            df[col] = None

    for _, row in df.iterrows():
        record = {col: (None if pd.isna(row.get(col)) else row.get(col)) for col in EXCEL_COLUMNS}
        fn = record.get("source_filename") or ""
        try:
            idx = int(record.get("item_index") or 0)
        except (TypeError, ValueError):
            idx = 0
        existing = _find_existing_row(ws, fn, idx) if fn and idx > 0 else None

        values = [record.get(col) for col in EXCEL_COLUMNS]
        if existing is not None:
            for col_idx, v in enumerate(values, start=1):
                ws.cell(row=existing, column=col_idx).value = v
            excel_row = existing
        else:
            ws.append(values)
            excel_row = ws.max_row

        if _needs_review(record):
            for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
                ws.cell(row=excel_row, column=col_idx).fill = HIGHLIGHT_FILL
        else:
            for col_idx in range(1, len(EXCEL_COLUMNS) + 1):
                ws.cell(row=excel_row, column=col_idx).fill = PatternFill(fill_type=None)

    _apply_column_widths(ws)
    _save_atomic(wb)
